# Read Excel sheet

from openpyxl import workbook,load_workbook
wb = load_workbook('C:\\Users\\ashok\\Downloads\\EmployeeDetails.xlsx')
ws = wb.active
print(ws)
print(ws['A1'].value)# to see column name
#status
ws['A1'].value = 'Status' # to change
wb.save('C:\\Users\\ashok\\Downloads\\EmployeeDetails.xlsx')
print(ws['A1'].value)
#Status
print(wb.sheetnames) # to check how no.of sheets present
#['Tablib Dataset']
print(wb['Tablib Dataset'])
#<Worksheet "Tablib Dataset">
wb.create_sheet('Test')# to create new sheet
#<Worksheet "Test">
print(wb.sheetnames)
#['Tablib Dataset', 'Test']


# create a Excel sheet
from openpyxl import workbook,load_workbook
wb = workbook()
ws = wb.active
ws.title = 'Excel Test'
ws.append('Name','Dept','Salary')
wb.save('Excel Practice')   

